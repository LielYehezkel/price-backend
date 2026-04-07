import csv
import hashlib
import io
import json
import re
import secrets
from collections import defaultdict

from io import BytesIO

from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, timezone
from typing import Annotated, Any
from urllib.parse import urlparse

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, func, or_
from sqlmodel import Session, select

from backend.config import settings
from backend.db import get_session
from backend.deps import get_current_user, require_shop_access
from backend.models import (
    Alert,
    ApiKey,
    CompetitorLink,
    DomainPriceApproval,
    DomainReviewQueueItem,
    Invite,
    PriceSnapshot,
    Product,
    ScanLog,
    Shop,
    ShopOwnershipTransfer,
    ShopMember,
    TrackedCompetitor,
    User,
    UserShopPreferences,
    WpConnectionToken,
    WpSetupToken,
    utcnow,
)
from backend.services.alert_prefs import (
    alert_allowed_by_prefs,
    get_or_create_user_shop_prefs,
    load_dismissed_recommendation_ids,
    save_dismissed_recommendation_ids,
)
from backend.services.domain_policy import domain_from_url, domain_is_live
from backend.services.domain_queue_repair import (
    ensure_domain_review_queue_item_for_competitor,
    repair_missing_domain_queue_for_shop,
)
from backend.services.extract import run_extraction_pipeline
from backend.services.fetch_html import fetch_html_sync
from backend.services.monitor_checks import run_competitor_check
from backend.services.system_config import resolve_public_api_base
from backend.services.wp_plugin_packager import build_plugin_zip_bytes
from backend.services.woo_sync import (
    effective_wc_price,
    fetch_wc_products,
    fetch_wc_products_by_ids,
    fetch_wc_store_currency,
    first_product_image_url,
)

router = APIRouter(prefix="/api/shops", tags=["shops"])


class ShopOut(BaseModel):
    id: int
    name: str
    check_interval_hours: int
    check_interval_minutes: int
    woocommerce_configured: bool
    woo_currency: str | None = None


class ShopCreate(BaseModel):
    name: str


class ShopPatch(BaseModel):
    name: str | None = None
    check_interval_hours: int | None = None
    check_interval_minutes: int | None = None


class ProductOut(BaseModel):
    id: int
    name: str
    sku: str | None
    permalink: str | None
    image_url: str | None = None
    category_name: str | None = None
    category_path: str | None = None
    regular_price: float | None
    competitors_count: int = 0
    shop_currency: str | None = None
    auto_pricing_enabled: bool = False
    auto_pricing_min_price: float | None = None
    auto_pricing_trigger_kind: str = "percent"
    auto_pricing_trigger_value: float | None = None
    auto_pricing_action_kind: str = "percent"
    auto_pricing_action_value: float | None = None
    auto_pricing_strategy: str = "reactive_down"


class ProductPageOut(BaseModel):
    items: list[ProductOut]
    total: int
    skip: int
    limit: int


class ProductAutoPricingPatch(BaseModel):
    auto_pricing_enabled: bool | None = None
    auto_pricing_min_price: float | None = None
    auto_pricing_trigger_kind: str | None = None
    auto_pricing_trigger_value: float | None = None
    auto_pricing_action_kind: str | None = None
    auto_pricing_action_value: float | None = None
    auto_pricing_strategy: str | None = None


class ProductCategoryRow(BaseModel):
    name: str
    count: int


class CompetitorOut(BaseModel):
    id: int
    url: str
    label: str | None
    domain: str
    display_name: str
    tracked_competitor_id: int | None = None
    last_price: float | None
    last_checked_at: datetime | None
    price_status: str = "live"  # live | processing


def _competitor_display(session: Session, c: CompetitorLink) -> tuple[str, str, int | None]:
    dom = domain_from_url(c.url)
    tid = getattr(c, "tracked_competitor_id", None)
    if tid:
        tc = session.get(TrackedCompetitor, tid)
        if tc:
            return dom, tc.display_name, tc.id
    lab = (c.label or "").strip()
    return dom, lab or dom, tid


def _competitor_to_out(session: Session, c: CompetitorLink) -> CompetitorOut:
    dom, disp, tid = _competitor_display(session, c)
    live = domain_is_live(session, dom)
    if live:
        return CompetitorOut(
            id=c.id,
            url=c.url,
            label=c.label,
            domain=dom,
            display_name=disp,
            tracked_competitor_id=tid,
            last_price=c.last_price,
            last_checked_at=c.last_checked_at,
            price_status="live",
        )
    return CompetitorOut(
        id=c.id,
        url=c.url,
        label=c.label,
        domain=dom,
        display_name=disp,
        tracked_competitor_id=tid,
        last_price=None,
        last_checked_at=c.last_checked_at,
        price_status="processing",
    )


class CompetitorCreate(BaseModel):
    url: str
    tracked_competitor_id: int | None = None
    competitor_name: str | None = None
    label: str | None = None  # תאימות לאחור; מועדף competitor_name


class BulkCompetitorsIn(BaseModel):
    urls_text: str
    label_prefix: str | None = None


def _resolve_tracked_competitor(
    session: Session,
    shop_id: int,
    url: str,
    tracked_competitor_id: int | None,
    competitor_name: str | None,
    legacy_label: str | None,
    *,
    allow_domain_fallback_name: bool = False,
) -> TrackedCompetitor:
    dom = domain_from_url(url.strip())
    if not dom:
        raise HTTPException(400, "כתובת לא תקינה")
    if tracked_competitor_id is not None:
        tc = session.get(TrackedCompetitor, tracked_competitor_id)
        if not tc or tc.shop_id != shop_id:
            raise HTTPException(400, "המתחרה שנבחר לא נמצא בחנות זו")
        if tc.domain != dom:
            raise HTTPException(
                400,
                "הדומיין בקישור חייב להתאים למתחרה שנבחר (אותו אתר).",
            )
        return tc
    existing = session.exec(
        select(TrackedCompetitor).where(
            TrackedCompetitor.shop_id == shop_id,
            TrackedCompetitor.domain == dom,
        ),
    ).first()
    if existing:
        return existing
    name = (competitor_name or legacy_label or "").strip()
    if not name and allow_domain_fallback_name:
        name = dom
    if not name:
        raise HTTPException(
            400,
            "דומיין חדש בחנות: יש לתת שם למתחרה (מזהה אנושי) — פעם אחת לכל אתר.",
        )
    tc = TrackedCompetitor(shop_id=shop_id, domain=dom, display_name=name[:240])
    session.add(tc)
    session.flush()
    return tc


class AlertOut(BaseModel):
    id: int
    message: str
    severity: str
    read: bool
    created_at: datetime
    kind: str = "general"


class UserShopPreferencesOut(BaseModel):
    notify_competitor_cheaper: bool
    notify_price_change: bool
    notify_auto_pricing: bool
    notify_sanity: bool


class UserShopPreferencesPatch(BaseModel):
    notify_competitor_cheaper: bool | None = None
    notify_price_change: bool | None = None
    notify_auto_pricing: bool | None = None
    notify_sanity: bool | None = None


class RecommendationsDismissIn(BaseModel):
    ids: list[str]


class MemberOut(BaseModel):
    user_id: int
    email: str
    role: str


class InviteCreate(BaseModel):
    email: str
    role: str = "member"


class ApiKeyOut(BaseModel):
    id: int
    name: str
    created_at: datetime
    prefix: str


class ApiKeyCreate(BaseModel):
    name: str


class WooConfig(BaseModel):
    site_url: str
    consumer_key: str
    consumer_secret: str


class OwnershipTransferCreateIn(BaseModel):
    target_email: str
    note: str | None = None


class OwnershipTransferOut(BaseModel):
    id: int
    shop_id: int
    shop_name: str
    from_user_id: int
    from_email: str
    to_user_id: int
    to_email: str
    status: str
    note: str | None
    created_at: str
    expires_at: str
    responded_at: str | None


def _ownership_transfer_to_out(session: Session, row: ShopOwnershipTransfer) -> OwnershipTransferOut:
    shop = session.get(Shop, row.shop_id)
    from_user = session.get(User, row.from_user_id)
    return OwnershipTransferOut(
        id=row.id or 0,
        shop_id=row.shop_id,
        shop_name=shop.name if shop else f"חנות #{row.shop_id}",
        from_user_id=row.from_user_id,
        from_email=from_user.email if from_user else f"user-{row.from_user_id}",
        to_user_id=row.to_user_id,
        to_email=row.to_email,
        status=row.status,
        note=row.note,
        created_at=row.created_at.isoformat(),
        expires_at=row.expires_at.isoformat(),
        responded_at=row.responded_at.isoformat() if row.responded_at else None,
    )


def _shop_to_out(s: Shop) -> ShopOut:
    mins = getattr(s, "check_interval_minutes", None)
    if mins is None or mins < 1:
        mins = max(1, int((s.check_interval_hours or 6) * 60))
    mins = max(1, min(int(mins), 60 * 24 * 14))
    return ShopOut(
        id=s.id,
        name=s.name,
        check_interval_hours=s.check_interval_hours,
        check_interval_minutes=mins,
        woocommerce_configured=bool(s.woo_site_url and s.woo_consumer_key),
        woo_currency=getattr(s, "woo_currency", None),
    )


def _product_to_out(session: Session, p: Product, shop: Shop) -> ProductOut:
    cc = session.exec(
        select(func.count(CompetitorLink.id)).where(CompetitorLink.product_id == p.id),
    ).first()
    return ProductOut(
        id=p.id,
        name=p.name,
        sku=p.sku,
        permalink=p.permalink,
        image_url=getattr(p, "image_url", None),
        category_name=getattr(p, "category_name", None),
        category_path=getattr(p, "category_path", None),
        regular_price=p.regular_price,
        competitors_count=int(cc or 0),
        shop_currency=getattr(shop, "woo_currency", None),
        auto_pricing_enabled=bool(getattr(p, "auto_pricing_enabled", False)),
        auto_pricing_min_price=getattr(p, "auto_pricing_min_price", None),
        auto_pricing_trigger_kind=getattr(p, "auto_pricing_trigger_kind", None) or "percent",
        auto_pricing_trigger_value=getattr(p, "auto_pricing_trigger_value", None),
        auto_pricing_action_kind=getattr(p, "auto_pricing_action_kind", None) or "percent",
        auto_pricing_action_value=getattr(p, "auto_pricing_action_value", None),
        auto_pricing_strategy=getattr(p, "auto_pricing_strategy", None) or "reactive_down",
    )


def _shops_for_user(session: Session, user: User) -> list[Shop]:
    mids = session.exec(select(ShopMember.shop_id).where(ShopMember.user_id == user.id)).all()
    conds = [Shop.owner_id == user.id]
    if mids:
        conds.append(Shop.id.in_(mids))
    return list(session.exec(select(Shop).where(or_(*conds))).all())


@router.get("", response_model=list[ShopOut])
def list_shops(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(get_current_user)]):
    shops = _shops_for_user(session, user)
    return [_shop_to_out(s) for s in shops]


@router.post("", response_model=ShopOut)
def create_shop(
    body: ShopCreate,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = Shop(name=body.name, owner_id=user.id)
    session.add(shop)
    session.commit()
    session.refresh(shop)
    session.add(ShopMember(shop_id=shop.id, user_id=user.id, role="owner"))
    session.commit()
    return _shop_to_out(shop)


@router.get("/{shop_id}", response_model=ShopOut)
def get_shop(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    s = session.get(Shop, shop_id)
    assert s
    return _shop_to_out(s)


@router.patch("/{shop_id}", response_model=ShopOut)
def patch_shop(
    shop_id: int,
    body: ShopPatch,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    if body.name is not None:
        shop.name = body.name
    if body.check_interval_minutes is not None:
        m = max(1, min(body.check_interval_minutes, 60 * 24 * 14))
        if m != getattr(shop, "check_interval_minutes", None):
            shop.last_scan_cycle_at = None
        shop.check_interval_minutes = m
        shop.check_interval_hours = max(1, (m + 59) // 60)
    elif body.check_interval_hours is not None:
        shop.check_interval_hours = max(1, body.check_interval_hours)
        new_m = max(1, shop.check_interval_hours * 60)
        if new_m != getattr(shop, "check_interval_minutes", None):
            shop.last_scan_cycle_at = None
        shop.check_interval_minutes = new_m
    session.add(shop)
    session.commit()
    session.refresh(shop)
    return _shop_to_out(shop)


@router.get("/{shop_id}/dashboard-stats")
def dashboard_stats(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    shop = session.get(Shop, shop_id)
    assert shop
    prefs = get_or_create_user_shop_prefs(session, user.id, shop_id)
    dismissed = load_dismissed_recommendation_ids(prefs)

    products = session.exec(select(Product).where(Product.shop_id == shop_id)).all()
    product_count = len(products)
    comps = session.exec(select(CompetitorLink).join(Product).where(Product.shop_id == shop_id)).all()
    competitor_links = len(comps)
    tracked_pids = {c.product_id for c in comps}
    products_with_competitors = len(tracked_pids)

    tc = session.exec(select(func.count(ScanLog.id)).where(ScanLog.shop_id == shop_id)).first()
    total_scans = int(tc or 0)

    mins = max(1, int(getattr(shop, "check_interval_minutes", None) or (shop.check_interval_hours or 6) * 60))
    since = utcnow() - timedelta(minutes=mins)
    sw = session.exec(
        select(func.count(ScanLog.id)).where(
            ScanLog.shop_id == shop_id,
            ScanLog.created_at >= since,
        ),
    ).first()
    scans_in_last_interval = int(sw or 0)

    losing = session.exec(
        select(func.count(ScanLog.id)).where(
            ScanLog.shop_id == shop_id,
            ScanLog.comparison == "you_expensive",
        ),
    ).first()
    losing_n = int(losing or 0)

    unread_rows = session.exec(
        select(Alert).where(Alert.shop_id == shop_id, Alert.read == False),  # noqa: E712
    ).all()
    unread_filtered = sum(1 for a in unread_rows if alert_allowed_by_prefs(a, prefs))

    rec_items: list[dict[str, str]] = []
    if product_count == 0:
        rec_items.append(
            {
                "id": "rec_sync_products",
                "text": "סנכרנו מוצרים מ־WooCommerce כדי להתחיל לעקוב אחרי מחירים.",
            },
        )
    if competitor_links == 0 and product_count > 0:
        rec_items.append(
            {
                "id": "rec_add_competitors",
                "text": "הוסיפו קישורי מתחרים למוצרים — כך תראו מי זול יותר בשוק.",
            },
        )
    if losing_n > 0:
        rec_items.append(
            {
                "id": "rec_losing_scans",
                "text": (
                    f"זוהו {losing_n} סריקות שבהן המתחרה הוצג כזול ממכם — כדאי לבדוק בלוגים ובמסך המוצרים."
                ),
            },
        )
    if unread_filtered > 0:
        rec_items.append(
            {
                "id": "rec_unread_alerts",
                "text": "יש התראות שלא סומנו כנקראו — מומלץ לעבור על מסך ההתראות.",
            },
        )
    rec_items = [r for r in rec_items if r["id"] not in dismissed]
    if not rec_items and "rec_all_ok" not in dismissed:
        rec_items.append({"id": "rec_all_ok", "text": "כרגע הכל נראה מצוין!"})

    # הערכת משך מחזור מלא: כל הקישורים ברצף (~20 שניות לממוצע גס לקישור)
    est_cycle = (competitor_links * 20) / 60.0 if competitor_links else 0.0

    return {
        "product_count": product_count,
        "products_with_competitors": products_with_competitors,
        "competitor_links": competitor_links,
        "check_interval_minutes": mins,
        "total_scans": total_scans,
        "scans_in_last_interval_window": scans_in_last_interval,
        "scans_expected_per_full_cycle": competitor_links,
        "worker_interval_seconds": 5,
        "queue_explanation": (
            "מחזור סריקה מלא לחנות: כשהמרווח שהגדרת (בדקות) עבר, כל קישורי המתחרה "
            "רצים מיד אחד אחרי השני — בלי המתנה של 15 שניות בין קישור לקישור."
        ),
        "estimated_full_queue_minutes_rounded": round(est_cycle, 1),
        "recommendations": rec_items,
    }


class SetupStepOut(BaseModel):
    id: str
    title: str
    description: str
    done: bool
    cta_label: str
    cta_path: str


class SetupChecklistOut(BaseModel):
    dismissed: bool
    percent_complete: int
    steps: list[SetupStepOut]


@router.get("/{shop_id}/setup-checklist", response_model=SetupChecklistOut)
def setup_checklist(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    shop = session.get(Shop, shop_id)
    assert shop
    dismissed = bool(getattr(shop, "setup_checklist_dismissed", False))
    woo_ok = bool(shop.woo_site_url and shop.woo_consumer_key and shop.woo_consumer_secret)
    products = session.exec(select(Product).where(Product.shop_id == shop_id)).all()
    product_count = len(products)
    comps = session.exec(select(CompetitorLink).join(Product).where(Product.shop_id == shop_id)).all()
    competitor_links = len(comps)
    mins = max(1, int(getattr(shop, "check_interval_minutes", None) or (shop.check_interval_hours or 6) * 60))
    interval_custom = mins != 360

    steps: list[SetupStepOut] = [
        SetupStepOut(
            id="woocommerce",
            title="חיבור WooCommerce",
            description="חברו את החנות כדי לסנכרן מוצרים ולעדכן מחירים אוטומטית.",
            done=woo_ok,
            cta_label="הגדרות חיבור",
            cta_path="settings",
        ),
        SetupStepOut(
            id="sync_products",
            title="סנכרון מוצרים",
            description="משכו את קטלוג המוצרים והמחירים מהחנות שלכם.",
            done=product_count > 0,
            cta_label="מוצרים וסנכרון",
            cta_path="products",
        ),
        SetupStepOut(
            id="competitors",
            title="מעקב מתחרים",
            description="הוסיפו קישור לעמוד מוצר אצל מתחרה — כדי להשוות מחירים ולנתח את השוק.",
            done=competitor_links > 0,
            cta_label="ניהול מתחרים",
            cta_path="products",
        ),
        SetupStepOut(
            id="scan_interval",
            title="התאמת קצב סריקה",
            description="הגדירו כל כמה זמן לרוץ על כל קישורי המתחרים (ברירת מחדל: 6 שעות).",
            done=interval_custom,
            cta_label="מרווח סריקה",
            cta_path="settings",
        ),
    ]
    n_done = sum(1 for s in steps if s.done)
    pct = int(round(100 * n_done / len(steps))) if steps else 100
    return SetupChecklistOut(dismissed=dismissed, percent_complete=pct, steps=steps)


@router.post("/{shop_id}/setup-checklist/dismiss")
def dismiss_setup_checklist(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    shop.setup_checklist_dismissed = True
    session.add(shop)
    session.commit()
    return {"ok": True}


class HealthIssueOut(BaseModel):
    code: str
    severity: str
    title: str
    detail: str


class AccountHealthOut(BaseModel):
    status: str
    score: int
    issues: list[HealthIssueOut]
    summary: str


class CompetitorCurrentSummaryOut(BaseModel):
    cheaper: int
    expensive: int
    tie: int
    compared: int


class CompetitorIntelRowOut(BaseModel):
    tracked_competitor_id: int | None
    competitor_name: str
    domain: str
    links_count: int
    current_cheaper: int
    current_expensive: int
    current_tie: int
    current_compared: int
    price_changes_in_period: int
    last_price_change_at: str | None


class CompetitorIntelligenceOut(BaseModel):
    period_days: int
    current_overall: CompetitorCurrentSummaryOut
    total_price_changes_in_period: int
    competitors: list[CompetitorIntelRowOut]


@router.get("/{shop_id}/account-health", response_model=AccountHealthOut)
def account_health(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    shop = session.get(Shop, shop_id)
    assert shop
    issues: list[HealthIssueOut] = []
    woo_ok = bool(shop.woo_site_url and shop.woo_consumer_key and shop.woo_consumer_secret)
    if not woo_ok:
        issues.append(
            HealthIssueOut(
                code="woo_disconnected",
                severity="warning",
                title="חיבור WooCommerce חסר",
                detail="ללא חיבור לא ניתן לסנכרן מוצרים לעדכן מחירים אוטומטית מהמערכת.",
            ),
        )
    product_count = session.exec(select(func.count(Product.id)).where(Product.shop_id == shop_id)).first() or 0
    if int(product_count) == 0 and woo_ok:
        issues.append(
            HealthIssueOut(
                code="no_products",
                severity="warning",
                title="אין מוצרים מסונכרנים",
                detail="הריצו סנכרון ממסך המוצרים כדי למשוך את הקטלוג.",
            ),
        )
    comp_n = session.exec(
        select(func.count(CompetitorLink.id)).join(Product).where(Product.shop_id == shop_id),
    ).first() or 0
    if int(comp_n) == 0 and int(product_count) > 0:
        issues.append(
            HealthIssueOut(
                code="no_competitors",
                severity="info",
                title="אין עדיין קישורי מתחרים",
                detail="הוסיפו קישורים כדי לקבל השוואות מחיר ותמחור חכם.",
            ),
        )
    pending_q = session.exec(
        select(func.count(DomainReviewQueueItem.id)).where(
            DomainReviewQueueItem.shop_id == shop_id,
            DomainReviewQueueItem.status == "pending",
        ),
    ).first() or 0
    if int(pending_q) > 0:
        issues.append(
            HealthIssueOut(
                code="domain_queue",
                severity="warning",
                title="תור ביקורת דומיינים",
                detail=f"{int(pending_q)} פריטים ממתינים לבדיקת סלקטור מחיר — הצוות יטפל בהמשך.",
            ),
        )
    domains_in_shop = {domain_from_url(c.url) for c in session.exec(select(CompetitorLink).join(Product).where(Product.shop_id == shop_id)).all()}
    domains_in_shop.discard("")
    pending_dom_count = 0
    for dom in domains_in_shop:
        row = session.get(DomainPriceApproval, dom)
        if row and row.status == "pending":
            pending_dom_count += 1
    if pending_dom_count > 0:
        issues.append(
            HealthIssueOut(
                code="domains_pending_approval",
                severity="info",
                title=f"{pending_dom_count} דומיינים ממתינים לאישור מחיר",
                detail="מחירי מתחרים יוצגו במלואם לאחר אישור הדומיין בפאנל הניהול.",
            ),
        )
    unread_bad = session.exec(
        select(func.count(Alert.id)).where(
            Alert.shop_id == shop_id,
            Alert.read == False,  # noqa: E712
            or_(Alert.severity == "hot", Alert.severity == "error"),
        ),
    ).first() or 0
    if int(unread_bad) > 0:
        issues.append(
            HealthIssueOut(
                code="critical_alerts",
                severity="warning",
                title="התראות הדורשות תשומת לב",
                detail=f"{int(unread_bad)} התראות לא נקראו עם סימון דחיפות.",
            ),
        )

    crit = sum(1 for i in issues if i.severity == "critical")
    warn = sum(1 for i in issues if i.severity == "warning")
    if crit > 0:
        status = "critical"
    elif warn > 0:
        status = "warning"
    else:
        status = "ok"
    score = max(0, min(100, 100 - 15 * warn - 40 * crit - 5 * max(0, len(issues) - warn - crit)))
    if status == "ok":
        summary = "החשבון במצב טוב — המעקב והחיבורים נראים תקינים."
    elif status == "warning":
        summary = "יש כמה נקודות לשיפור; מומלץ לעבור על הרשימה למטה."
    else:
        summary = "זוהו בעיות שדורשות טיפול — בדקו את ההתראות והגדרות."

    return AccountHealthOut(status=status, score=score, issues=issues, summary=summary)


@router.get("/{shop_id}/competitors/intelligence", response_model=CompetitorIntelligenceOut)
def competitors_intelligence(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    days: int = Query(30, ge=1, le=365),
):
    """תמונת תחרות: זול/יקר/שווה כרגע + קצב שינויי מחיר בתקופה."""
    require_shop_access(session, user, shop_id)
    comps = session.exec(
        select(CompetitorLink)
        .join(Product, CompetitorLink.product_id == Product.id)
        .where(Product.shop_id == shop_id),
    ).all()
    if not comps:
        return CompetitorIntelligenceOut(
            period_days=days,
            current_overall=CompetitorCurrentSummaryOut(cheaper=0, expensive=0, tie=0, compared=0),
            total_price_changes_in_period=0,
            competitors=[],
        )

    # latest scan per competitor_link = מצב נוכחי מול אותו מתחרה.
    latest_by_link: dict[int, ScanLog] = {}
    for c in comps:
        lg = session.exec(
            select(ScanLog)
            .where(ScanLog.competitor_link_id == c.id)
            .order_by(ScanLog.id.desc())
            .limit(1),
        ).first()
        if lg:
            latest_by_link[c.id] = lg

    since = utcnow() - timedelta(days=days)
    change_logs = session.exec(
        select(ScanLog).where(
            ScanLog.shop_id == shop_id,
            ScanLog.created_at >= since,
            ScanLog.price_changed == True,  # noqa: E712
        ),
    ).all()
    changes_by_domain: dict[str, int] = defaultdict(int)
    last_change_at_by_domain: dict[str, datetime] = {}
    for lg in change_logs:
        dom = (lg.competitor_domain or "").strip().lower()
        if not dom:
            continue
        changes_by_domain[dom] += 1
        prev = last_change_at_by_domain.get(dom)
        if prev is None or lg.created_at > prev:
            last_change_at_by_domain[dom] = lg.created_at

    tracked_cache: dict[int, TrackedCompetitor] = {}
    grouped: dict[str, CompetitorIntelRowOut] = {}
    for c in comps:
        dom = domain_from_url(c.url)
        tid = getattr(c, "tracked_competitor_id", None)
        if tid and tid not in tracked_cache:
            tc = session.get(TrackedCompetitor, tid)
            if tc:
                tracked_cache[tid] = tc
        name = tracked_cache[tid].display_name if tid and tid in tracked_cache else ((c.label or "").strip() or dom)
        key = f"{tid or 0}:{dom}"
        row = grouped.get(key)
        if not row:
            row = CompetitorIntelRowOut(
                tracked_competitor_id=tid,
                competitor_name=name,
                domain=dom,
                links_count=0,
                current_cheaper=0,
                current_expensive=0,
                current_tie=0,
                current_compared=0,
                price_changes_in_period=0,
                last_price_change_at=None,
            )
            grouped[key] = row
        row.links_count += 1

        lg = latest_by_link.get(c.id)
        if lg and lg.comparison in {"you_cheaper", "you_expensive", "tie"}:
            row.current_compared += 1
            if lg.comparison == "you_cheaper":
                row.current_cheaper += 1
            elif lg.comparison == "you_expensive":
                row.current_expensive += 1
            elif lg.comparison == "tie":
                row.current_tie += 1

        row.price_changes_in_period += int(changes_by_domain.get(dom, 0))
        dt = last_change_at_by_domain.get(dom)
        if dt:
            row.last_price_change_at = dt.isoformat()

    competitors = sorted(
        grouped.values(),
        key=lambda r: (r.current_compared, r.current_expensive, r.price_changes_in_period),
        reverse=True,
    )
    overall = CompetitorCurrentSummaryOut(
        cheaper=sum(r.current_cheaper for r in competitors),
        expensive=sum(r.current_expensive for r in competitors),
        tie=sum(r.current_tie for r in competitors),
        compared=sum(r.current_compared for r in competitors),
    )
    return CompetitorIntelligenceOut(
        period_days=days,
        current_overall=overall,
        total_price_changes_in_period=sum(r.price_changes_in_period for r in competitors),
        competitors=competitors,
    )


@router.get("/{shop_id}/reports/weekly.csv")
def weekly_report_csv(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    shop = session.get(Shop, shop_id)
    assert shop
    since = utcnow() - timedelta(days=7)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["דוח שבועי — מעקב מחירים"])
    w.writerow(["חנות", shop.name, "מטבע", shop.woo_currency or ""])
    w.writerow(["תקופה", f"7 ימים אחרונים עד {utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
    w.writerow([])
    scan_total = session.exec(
        select(func.count(ScanLog.id)).where(ScanLog.shop_id == shop_id, ScanLog.created_at >= since),
    ).first() or 0
    w.writerow(["סה״כ סריקות מתועדות (7 ימים)", int(scan_total)])
    for cmp_key, label_he in [
        ("you_cheaper", "פעמים שהמחיר שלכם נמוך יותר"),
        ("you_expensive", "פעמים שהמתחרה הוצג כזול ממכם"),
        ("tie", "אותו מחיר"),
        ("unknown", "לא ניתן להשוות"),
    ]:
        n = session.exec(
            select(func.count(ScanLog.id)).where(
                ScanLog.shop_id == shop_id,
                ScanLog.created_at >= since,
                ScanLog.comparison == cmp_key,
            ),
        ).first() or 0
        w.writerow([label_he, int(n)])
    price_changes = session.exec(
        select(func.count(ScanLog.id)).where(
            ScanLog.shop_id == shop_id,
            ScanLog.created_at >= since,
            ScanLog.price_changed == True,  # noqa: E712
        ),
    ).first() or 0
    w.writerow(["שינויי מחיר מתחרה שזוהו", int(price_changes)])
    w.writerow([])
    alerts_n = session.exec(
        select(func.count(Alert.id)).where(Alert.shop_id == shop_id, Alert.created_at >= since),
    ).first() or 0
    w.writerow(["התראות שנוצרו בתקופה", int(alerts_n)])
    tracked_n = session.exec(select(func.count(TrackedCompetitor.id)).where(TrackedCompetitor.shop_id == shop_id)).first() or 0
    w.writerow(["מתחרים מזוהים (דומיינים)", int(tracked_n)])
    products_n = session.exec(select(func.count(Product.id)).where(Product.shop_id == shop_id)).first() or 0
    w.writerow(["מוצרים בחנות", int(products_n)])
    comp_links = session.exec(
        select(func.count(CompetitorLink.id)).join(Product).where(Product.shop_id == shop_id),
    ).first() or 0
    w.writerow(["קישורי מתחרים פעילים", int(comp_links)])
    data = buf.getvalue().encode("utf-8-sig")
    fname = f"weekly-report-shop-{shop_id}.csv"
    return Response(
        content=data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{shop_id}/analytics/sales-insights")
def sales_insights(
    shop_id: int,
    background_tasks: BackgroundTasks,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    days: int = Query(90, ge=7, le=365),
    force_refresh: bool = Query(False),
):
    """הכנסות והתפלגות מחירים מ-WooCommerce. מטמון מהיר + רענון ברקע כל ~45 דק׳."""
    require_shop_access(session, user, shop_id)
    from backend.services import woo_analytics as wa

    days = max(7, min(days, 365))
    now = utcnow()

    if force_refresh:
        try:
            data = wa.compute_sales_insights(session, shop_id, days)
            if data.get("ok"):
                wa.save_sales_insights_cache(session, shop_id, days, data)
            return wa.attach_cache_meta(data, fresh=True, computed_at=now)
        except Exception as ex:
            return {
                "ok": False,
                "error": "sales_insights_failed",
                "message_he": f"דוח מכירות נכשל כרגע: {ex!s}",
            }

    row = wa.get_sales_insights_cache_row(session, shop_id, days)
    if row is not None:
        try:
            cached = json.loads(row.payload_json)
        except json.JSONDecodeError:
            cached = None
        if isinstance(cached, dict) and cached.get("ok") is not None:
            if wa.is_sales_cache_fresh(row, now):
                return wa.attach_cache_meta(cached, fresh=True, stale=False, computed_at=row.updated_at)
            background_tasks.add_task(wa.refresh_sales_insights_cache_task, shop_id, days)
            return wa.attach_cache_meta(cached, fresh=False, stale=True, computed_at=row.updated_at)

    try:
        data = wa.compute_sales_insights(session, shop_id, days)
        if data.get("ok"):
            wa.save_sales_insights_cache(session, shop_id, days, data)
        return wa.attach_cache_meta(data, fresh=True, computed_at=now)
    except Exception as ex:
        return {
            "ok": False,
            "error": "sales_insights_failed",
            "message_he": f"דוח מכירות נכשל כרגע: {ex!s}",
        }


class ScanLogOut(BaseModel):
    id: int
    created_at: datetime
    product_name: str
    competitor_domain: str
    our_price: float | None
    competitor_price: float | None
    previous_competitor_price: float | None
    price_changed: bool
    comparison: str
    comparison_label: str


class ScanLogsPage(BaseModel):
    items: list[ScanLogOut]
    total: int


@router.get("/{shop_id}/scan-logs", response_model=ScanLogsPage)
def scan_logs(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
):
    require_shop_access(session, user, shop_id)
    conds = [
        ScanLog.shop_id == shop_id,
        ScanLog.comparison != "pending_review",
    ]
    if not user.is_admin:
        conds.append(ScanLog.comparison != "sanity_failed")
    base = select(ScanLog).where(and_(*conds)).order_by(ScanLog.created_at.desc())
    total = session.exec(select(func.count(ScanLog.id)).where(and_(*conds))).first()
    logs = session.exec(base.offset(skip).limit(limit)).all()
    labels = {
        "you_cheaper": "אתה זול יותר מהמתחרה",
        "tie": "אותו מחיר (בערך)",
        "you_expensive": "המתחרה זול ממך",
        "unknown": "לא ניתן להשוות",
        "sanity_failed": "נדחה — מחיר לא סביר (סף אמינות)",
    }
    out: list[ScanLogOut] = []
    for lg in logs:
        out.append(
            ScanLogOut(
                id=lg.id,
                created_at=lg.created_at,
                product_name=lg.product_name,
                competitor_domain=lg.competitor_domain,
                our_price=lg.our_price,
                competitor_price=lg.competitor_price,
                previous_competitor_price=lg.previous_competitor_price,
                price_changed=lg.price_changed,
                comparison=lg.comparison,
                comparison_label=labels.get(lg.comparison, lg.comparison),
            ),
        )
    return ScanLogsPage(items=out, total=int(total or 0))


@router.post("/{shop_id}/sync")
def sync_shop(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    if not shop.woo_site_url or not shop.woo_consumer_key or not shop.woo_consumer_secret:
        raise HTTPException(400, "יש לשמור פרטי WooCommerce בהגדרות")
    try:
        rows = fetch_wc_products(shop.woo_site_url, shop.woo_consumer_key, shop.woo_consumer_secret)
        cur = fetch_wc_store_currency(shop.woo_site_url, shop.woo_consumer_key, shop.woo_consumer_secret)
    except Exception as ex:
        raise HTTPException(
            400,
            f"סנכרון מול WooCommerce נכשל: {ex!s}. בדקו כתובת אתר, מפתחות API, SSL והרשאות read_write.",
        ) from ex
    if cur:
        shop.woo_currency = cur
        session.add(shop)
    n = 0
    for r in rows:
        if not isinstance(r, dict):
            continue
        wid = r.get("id")
        name = r.get("name") or "—"
        sku = r.get("sku")
        link = r.get("permalink")
        # מחיר מכירה בפועל: אם יש מבצע Woo מחזיר אותו ב-price.
        # fallback ל-sale_price ואז ל-regular_price.
        price = effective_wc_price(r)
        img = first_product_image_url(r)
        cats_raw = r.get("categories")
        cat_names: list[str] = []
        if isinstance(cats_raw, list):
            for c in cats_raw:
                if isinstance(c, dict):
                    nm = str(c.get("name") or "").strip()
                    if nm:
                        cat_names.append(nm)
        category_name = cat_names[0] if cat_names else None
        category_path = " > ".join(cat_names) if cat_names else None
        existing = session.exec(
            select(Product).where(Product.shop_id == shop.id, Product.woo_product_id == wid)
        ).first()
        if existing:
            existing.name = str(name)
            existing.sku = str(sku) if sku else None
            existing.permalink = str(link) if link else None
            old_price = existing.regular_price
            if old_price is None or price is None:
                changed = old_price != price
            else:
                changed = abs(float(old_price) - float(price)) > 0.005
            if changed:
                existing.regular_price = price
            existing.last_price_sync_at = utcnow()
            existing.image_url = img
            existing.category_name = category_name
            existing.category_path = category_path
            session.add(existing)
        else:
            session.add(
                Product(
                    shop_id=shop.id,
                    woo_product_id=int(wid) if wid is not None else None,
                    name=str(name),
                    sku=str(sku) if sku else None,
                    permalink=str(link) if link else None,
                    image_url=img,
                    category_name=category_name,
                    category_path=category_path,
                    regular_price=price,
                    last_price_sync_at=utcnow(),
                )
            )
        n += 1
    session.commit()
    return {"synced": n}


@router.post("/{shop_id}/refresh-prices")
def refresh_shop_prices(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    """
    מעדכן מחירי מכירה בפועל (ומבצעים) רק למוצרים שכבר קיימים במערכת —
    בלי לייבא מחדש את כל הקטלוג ומבלי להוסיף מוצרים חדשים מהחנות.
    """
    shop = require_shop_access(session, user, shop_id)
    if not shop.woo_site_url or not shop.woo_consumer_key or not shop.woo_consumer_secret:
        raise HTTPException(400, "יש לשמור פרטי WooCommerce בהגדרות")

    rows = list(
        session.exec(
            select(Product).where(
                Product.shop_id == shop_id,
                Product.woo_product_id.is_not(None),
            ),
        ).all(),
    )
    woo_ids = sorted({int(p.woo_product_id) for p in rows if p.woo_product_id is not None})
    if not woo_ids:
        return {"checked": 0, "updated": 0, "missing_in_woo": 0}

    try:
        wc_by_id = fetch_wc_products_by_ids(
            shop.woo_site_url,
            shop.woo_consumer_key,
            shop.woo_consumer_secret,
            woo_ids,
        )
    except Exception as ex:
        raise HTTPException(
            400,
            f"רענון מחירים מול WooCommerce נכשל: {ex!s}. בדקו חיבור, מפתחות API והרשאות.",
        ) from ex

    cur = fetch_wc_store_currency(shop.woo_site_url, shop.woo_consumer_key, shop.woo_consumer_secret)
    if cur:
        shop.woo_currency = cur
        session.add(shop)

    now = utcnow()
    updated = 0
    checked = len(rows)
    missing = 0

    for p in rows:
        if p.woo_product_id is None:
            continue
        wid = int(p.woo_product_id)
        row = wc_by_id.get(wid)
        if not row:
            missing += 1
            continue
        new_price = effective_wc_price(row)
        old_price = p.regular_price
        if old_price is None or new_price is None:
            price_changed = old_price != new_price
        else:
            price_changed = abs(float(old_price) - float(new_price)) > 0.005
        if price_changed:
            p.regular_price = new_price
            updated += 1
        p.last_price_sync_at = now
        session.add(p)

    session.commit()
    return {
        "checked": checked,
        "updated": updated,
        "missing_in_woo": missing,
    }


@router.get("/{shop_id}/products", response_model=ProductPageOut)
def list_products(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    q: str | None = None,
    category: str | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(80, ge=1, le=300),
):
    shop = require_shop_access(session, user, shop_id)
    stmt = select(Product).where(Product.shop_id == shop_id)
    cat = (category or "").strip()
    if cat and cat != "__all__":
        if cat == "__uncategorized__":
            stmt = stmt.where(
                or_(Product.category_name.is_(None), Product.category_name == ""),
            )
        else:
            stmt = stmt.where(Product.category_name == cat)
    if q:
        qq = q.strip()
        if qq:
            like = f"%{qq}%"
            stmt = stmt.where(or_(Product.name.like(like), Product.sku.like(like)))

    total = session.exec(select(func.count()).select_from(stmt.subquery())).first() or 0
    rows = session.exec(stmt.order_by(Product.id).offset(skip).limit(limit)).all()
    out = [_product_to_out(session, p, shop) for p in rows]
    return ProductPageOut(items=out, total=int(total), skip=skip, limit=limit)


@router.get("/{shop_id}/products/categories", response_model=list[ProductCategoryRow])
def list_product_categories(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    rows = session.exec(
        select(Product.category_name, func.count(Product.id))
        .where(Product.shop_id == shop_id, Product.category_name.is_not(None))
        .group_by(Product.category_name)
        .order_by(func.count(Product.id).desc(), Product.category_name.asc()),
    ).all()
    out: list[ProductCategoryRow] = []
    for name, count in rows:
        nm = (name or "").strip()
        if not nm:
            continue
        out.append(ProductCategoryRow(name=nm, count=int(count or 0)))
    return out


@router.patch("/{shop_id}/products/{product_id}", response_model=ProductOut)
def patch_product_auto_pricing(
    shop_id: int,
    product_id: int,
    body: ProductAutoPricingPatch,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    p = session.get(Product, product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "מוצר לא נמצא")
    if body.auto_pricing_enabled is not None:
        p.auto_pricing_enabled = body.auto_pricing_enabled
    if body.auto_pricing_min_price is not None:
        p.auto_pricing_min_price = body.auto_pricing_min_price
    if body.auto_pricing_trigger_kind is not None:
        if body.auto_pricing_trigger_kind not in ("percent", "amount"):
            raise HTTPException(400, "trigger_kind חייב להיות percent או amount")
        p.auto_pricing_trigger_kind = body.auto_pricing_trigger_kind
    if body.auto_pricing_trigger_value is not None:
        p.auto_pricing_trigger_value = body.auto_pricing_trigger_value
    if body.auto_pricing_action_kind is not None:
        if body.auto_pricing_action_kind not in ("percent", "amount"):
            raise HTTPException(400, "action_kind חייב להיות percent או amount")
        p.auto_pricing_action_kind = body.auto_pricing_action_kind
    if body.auto_pricing_action_value is not None:
        p.auto_pricing_action_value = body.auto_pricing_action_value
    if body.auto_pricing_strategy is not None:
        if body.auto_pricing_strategy not in ("reactive_down", "smart_anchor"):
            raise HTTPException(400, "אסטרטגיית תמחור לא תקינה")
        p.auto_pricing_strategy = body.auto_pricing_strategy

    if p.auto_pricing_enabled:
        if p.auto_pricing_min_price is None or p.auto_pricing_min_price <= 0:
            raise HTTPException(400, "חובה מחיר מינימום חיובי כשמפעילים תמחור אוטומטי")
        if p.auto_pricing_action_value is None or p.auto_pricing_action_value < 0:
            raise HTTPException(400, "חובה ערך פעולה (לא שלילי)")
        strat = getattr(p, "auto_pricing_strategy", None) or "reactive_down"
        if strat == "reactive_down":
            if p.auto_pricing_trigger_value is None or p.auto_pricing_trigger_value < 0:
                raise HTTPException(400, "במצב 'רק הורדה' חובה ערך תנאי (לא שלילי)")
        else:
            if p.auto_pricing_trigger_value is None:
                p.auto_pricing_trigger_value = 0.0

    session.add(p)
    session.commit()
    session.refresh(p)
    return _product_to_out(session, p, shop)


@router.get("/{shop_id}/competitor-labels")
def competitor_label_suggestions(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    tracked = session.exec(
        select(TrackedCompetitor).where(TrackedCompetitor.shop_id == shop_id),
    ).all()
    seen: set[str] = {t.display_name.strip().lower() for t in tracked if t.display_name}
    labels: list[str] = [t.display_name for t in tracked if (t.display_name or "").strip()]
    comps = session.exec(
        select(CompetitorLink).join(Product).where(Product.shop_id == shop_id),
    ).all()
    for c in comps:
        lab = (c.label or "").strip()
        if not lab or lab.lower() in seen:
            continue
        seen.add(lab.lower())
        labels.append(lab)
    labels.sort(key=str.lower)
    return {"labels": labels}


class TrackedCompetitorRow(BaseModel):
    id: int
    domain: str
    display_name: str
    links_count: int


class TrackedCompetitorPatch(BaseModel):
    display_name: str


@router.get("/{shop_id}/tracked-competitors", response_model=list[TrackedCompetitorRow])
def list_tracked_competitors(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    rows = session.exec(
        select(TrackedCompetitor).where(TrackedCompetitor.shop_id == shop_id).order_by(TrackedCompetitor.domain),
    ).all()
    out: list[TrackedCompetitorRow] = []
    for t in rows:
        n = session.exec(
            select(func.count(CompetitorLink.id)).where(CompetitorLink.tracked_competitor_id == t.id),
        ).first()
        out.append(
            TrackedCompetitorRow(
                id=t.id,
                domain=t.domain,
                display_name=t.display_name,
                links_count=int(n or 0),
            ),
        )
    return out


@router.patch("/{shop_id}/tracked-competitors/{tracked_id}", response_model=TrackedCompetitorRow)
def patch_tracked_competitor(
    shop_id: int,
    tracked_id: int,
    body: TrackedCompetitorPatch,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    t = session.get(TrackedCompetitor, tracked_id)
    if not t or t.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    name = body.display_name.strip()
    if not name:
        raise HTTPException(400, "שם תצוגה לא יכול להיות ריק")
    t.display_name = name[:240]
    session.add(t)
    for c in session.exec(
        select(CompetitorLink).where(CompetitorLink.tracked_competitor_id == t.id),
    ).all():
        c.label = t.display_name
        session.add(c)
    session.commit()
    session.refresh(t)
    n = session.exec(
        select(func.count(CompetitorLink.id)).where(CompetitorLink.tracked_competitor_id == t.id),
    ).first()
    return TrackedCompetitorRow(
        id=t.id,
        domain=t.domain,
        display_name=t.display_name,
        links_count=int(n or 0),
    )


@router.get("/{shop_id}/products/{product_id}/competitors", response_model=list[CompetitorOut])
def list_competitors(
    shop_id: int,
    product_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = session.get(Product, product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "מוצר לא נמצא")
    repair_missing_domain_queue_for_shop(session, shop_id)
    comps = session.exec(select(CompetitorLink).where(CompetitorLink.product_id == product_id)).all()
    return [_competitor_to_out(session, c) for c in comps]


@router.post("/{shop_id}/products/{product_id}/competitors", response_model=CompetitorOut)
def add_competitor(
    shop_id: int,
    product_id: int,
    body: CompetitorCreate,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = session.get(Product, product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "מוצר לא נמצא")
    url_s = body.url.strip()
    tc = _resolve_tracked_competitor(
        session,
        shop_id,
        url_s,
        body.tracked_competitor_id,
        body.competitor_name,
        body.label,
        allow_domain_fallback_name=False,
    )
    disp = tc.display_name
    c = CompetitorLink(
        product_id=product_id,
        url=url_s,
        label=disp,
        tracked_competitor_id=tc.id,
    )
    session.add(c)
    session.commit()
    session.refresh(c)
    ensure_domain_review_queue_item_for_competitor(session, shop_id, c.id, try_fetch=True)
    session.commit()
    return _competitor_to_out(session, c)


def _normalize_competitor_url(line: str) -> str | None:
    raw = line.strip()
    if not raw:
        return None
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    p = urlparse(raw)
    if not p.netloc:
        return None
    return raw


@router.post("/{shop_id}/products/{product_id}/competitors/bulk")
def bulk_add_competitors(
    shop_id: int,
    product_id: int,
    body: BulkCompetitorsIn,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = session.get(Product, product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "מוצר לא נמצא")
    lines = [ln for ln in body.urls_text.replace("\r", "\n").split("\n") if ln.strip()]
    added = 0
    for i, line in enumerate(lines):
        url = _normalize_competitor_url(line)
        if not url:
            continue
        exists = session.exec(
            select(CompetitorLink).where(
                CompetitorLink.product_id == product_id,
                CompetitorLink.url == url,
            ),
        ).first()
        if exists:
            continue
        legacy = None
        if body.label_prefix:
            legacy = f"{body.label_prefix.strip()} {i + 1}"
        tc = _resolve_tracked_competitor(
            session,
            shop_id,
            url,
            None,
            None,
            legacy,
            allow_domain_fallback_name=True,
        )
        c = CompetitorLink(
            product_id=product_id,
            url=url,
            label=tc.display_name,
            tracked_competitor_id=tc.id,
        )
        session.add(c)
        session.flush()
        ensure_domain_review_queue_item_for_competitor(session, shop_id, c.id, try_fetch=False)
        added += 1
    session.commit()
    return {"added": added, "lines_processed": len(lines)}


@router.post("/{shop_id}/competitors/{competitor_id}/check")
def check_competitor(
    shop_id: int,
    competitor_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    c = session.get(CompetitorLink, competitor_id)
    if not c:
        raise HTTPException(404, "לא נמצא")
    p = session.get(Product, c.product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    result = run_competitor_check(session, competitor_id)
    return {
        "price": result.price if result.published else None,
        "currency": result.currency if result.published else None,
        "price_status": "live" if result.published else "processing",
    }


class ReportPriceIssueIn(BaseModel):
    note: str | None = None


@router.post("/{shop_id}/competitors/{competitor_id}/report-price-issue")
def report_competitor_price_issue(
    shop_id: int,
    competitor_id: int,
    body: ReportPriceIssueIn,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    """דיווח לקוח: מחיר נראה שגוי — נכנס לתור ביקורת דומיין בפאנל ניהול."""
    require_shop_access(session, user, shop_id)
    c = session.get(CompetitorLink, competitor_id)
    if not c:
        raise HTTPException(404, "לא נמצא")
    p = session.get(Product, c.product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    domain = domain_from_url(c.url)
    if not domain:
        raise HTTPException(400, "כתובת לא תקינה")
    html = fetch_html_sync(c.url)
    result = run_extraction_pipeline(html)
    price = result.get("price")
    cur = result.get("currency")
    candidates = result.get("candidates") or []
    sug = candidates[0].get("selector") if candidates and isinstance(candidates[0], dict) else None

    if not domain_is_live(session, domain):
        dpa = session.get(DomainPriceApproval, domain)
        if not dpa:
            dpa = DomainPriceApproval(domain=domain)
        dpa.status = "pending"
        dpa.sample_url = c.url
        dpa.pending_price = price
        dpa.pending_currency = cur
        dpa.candidates_json = json.dumps(candidates[:40], ensure_ascii=False)
        dpa.suggested_selector = sug
        dpa.updated_at = utcnow()
        session.add(dpa)

    row = DomainReviewQueueItem(
        domain=domain,
        competitor_link_id=c.id,
        shop_id=shop_id,
        product_name=p.name or "",
        sample_url=c.url,
        pending_price=price,
        pending_currency=cur,
        candidates_json=json.dumps(candidates[:40], ensure_ascii=False),
        suggested_selector=sug,
        source="user_report",
        reporter_note=(body.note or "").strip() or None,
        status="pending",
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return {"ok": True, "queue_item_id": row.id, "domain": domain}


class SnapshotOut(BaseModel):
    id: int
    price: float | None
    currency: str | None
    fetched_at: datetime


@router.get("/{shop_id}/competitors/{competitor_id}/snapshots", response_model=list[SnapshotOut])
def list_snapshots(
    shop_id: int,
    competitor_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    c = session.get(CompetitorLink, competitor_id)
    if not c:
        raise HTTPException(404, "לא נמצא")
    p = session.get(Product, c.product_id)
    if not p or p.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    if not domain_is_live(session, domain_from_url(c.url)):
        return []
    snaps = session.exec(
        select(PriceSnapshot)
        .where(PriceSnapshot.competitor_link_id == competitor_id)
        .order_by(PriceSnapshot.fetched_at.desc())
    ).all()
    return [
        SnapshotOut(id=s.id, price=s.price, currency=s.currency, fetched_at=s.fetched_at) for s in snaps
    ]


@router.get("/{shop_id}/alerts", response_model=list[AlertOut])
def list_alerts(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    unread_only: bool = Query(False),
):
    require_shop_access(session, user, shop_id)
    prefs = get_or_create_user_shop_prefs(session, user.id, shop_id)
    q = select(Alert).where(Alert.shop_id == shop_id)
    if unread_only:
        q = q.where(Alert.read == False)  # noqa: E712
    alerts = session.exec(q.order_by(Alert.created_at.desc())).all()
    alerts = [a for a in alerts if alert_allowed_by_prefs(a, prefs)]
    return [
        AlertOut(
            id=a.id,
            message=a.message,
            severity=a.severity,
            read=a.read,
            created_at=a.created_at,
            kind=getattr(a, "kind", None) or "general",
        )
        for a in alerts
    ]


@router.get("/{shop_id}/notification-preferences", response_model=UserShopPreferencesOut)
def get_notification_preferences(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = get_or_create_user_shop_prefs(session, user.id, shop_id)
    return UserShopPreferencesOut(
        notify_competitor_cheaper=p.notify_competitor_cheaper,
        notify_price_change=p.notify_price_change,
        notify_auto_pricing=p.notify_auto_pricing,
        notify_sanity=p.notify_sanity,
    )


@router.patch("/{shop_id}/notification-preferences", response_model=UserShopPreferencesOut)
def patch_notification_preferences(
    shop_id: int,
    body: UserShopPreferencesPatch,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = get_or_create_user_shop_prefs(session, user.id, shop_id)
    if body.notify_competitor_cheaper is not None:
        p.notify_competitor_cheaper = body.notify_competitor_cheaper
    if body.notify_price_change is not None:
        p.notify_price_change = body.notify_price_change
    if body.notify_auto_pricing is not None:
        p.notify_auto_pricing = body.notify_auto_pricing
    if body.notify_sanity is not None:
        p.notify_sanity = body.notify_sanity
    p.updated_at = utcnow()
    session.add(p)
    session.commit()
    session.refresh(p)
    return UserShopPreferencesOut(
        notify_competitor_cheaper=p.notify_competitor_cheaper,
        notify_price_change=p.notify_price_change,
        notify_auto_pricing=p.notify_auto_pricing,
        notify_sanity=p.notify_sanity,
    )


@router.post("/{shop_id}/recommendations/dismiss")
def dismiss_recommendations(
    shop_id: int,
    body: RecommendationsDismissIn,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    p = get_or_create_user_shop_prefs(session, user.id, shop_id)
    cur = load_dismissed_recommendation_ids(p)
    for i in body.ids:
        if i.strip():
            cur.add(i.strip())
    save_dismissed_recommendation_ids(session, p, cur)
    return {"ok": True, "dismissed_count": len(cur)}


@router.post("/{shop_id}/alerts/{alert_id}/read")
def read_alert(
    shop_id: int,
    alert_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    a = session.get(Alert, alert_id)
    if not a or a.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    a.read = True
    session.add(a)
    session.commit()
    return {"ok": True}


@router.post("/{shop_id}/alerts/read-all")
def read_all_alerts(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    alerts = session.exec(select(Alert).where(Alert.shop_id == shop_id, Alert.read == False)).all()  # noqa: E712
    for a in alerts:
        a.read = True
        session.add(a)
    session.commit()
    return {"ok": True}


class SeriesOut(BaseModel):
    points: list[dict[str, Any]]


def _hour_bucket_utc(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        d = dt.replace(tzinfo=timezone.utc)
    else:
        d = dt.astimezone(timezone.utc)
    return d.replace(minute=0, second=0, microsecond=0)


@router.get("/{shop_id}/analytics/price-series", response_model=SeriesOut)
def price_series(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    product_id: int | None = None,
    competitor_id: int | None = None,
    aggregate: str | None = Query(
        None,
        description="hourly_min: מחיר המתחרה הנמוך ביותר בכל שעה (מומלץ לדשבורד). ברירת מחדל: כל הדגימות הגולמיות.",
    ),
):
    require_shop_access(session, user, shop_id)
    comps = session.exec(
        select(CompetitorLink)
        .join(Product)
        .where(Product.shop_id == shop_id)
    ).all()
    if competitor_id is not None:
        comps = [c for c in comps if c.id == competitor_id]
    if product_id is not None:
        comps = [c for c in comps if c.product_id == product_id]
    raw_rows: list[tuple[datetime, float]] = []
    for c in comps:
        if not domain_is_live(session, domain_from_url(c.url)):
            continue
        snaps = session.exec(
            select(PriceSnapshot)
            .where(PriceSnapshot.competitor_link_id == c.id)
            .order_by(PriceSnapshot.fetched_at.asc())
        ).all()
        for s in snaps:
            if s.price is not None:
                raw_rows.append((s.fetched_at, float(s.price)))

    if aggregate == "hourly_min":
        buckets: dict[datetime, list[float]] = defaultdict(list)
        for t, price in raw_rows:
            buckets[_hour_bucket_utc(t)].append(price)
        points: list[dict[str, Any]] = []
        for b in sorted(buckets.keys()):
            vals = buckets[b]
            points.append(
                {
                    "t": b.isoformat(),
                    "price": min(vals),
                    "samples": len(vals),
                },
            )
        return SeriesOut(points=points[-400:])

    points = [{"t": t.isoformat(), "price": p} for t, p in raw_rows]
    points.sort(key=lambda x: x["t"])
    return SeriesOut(points=points[-400:])


@router.get("/{shop_id}/analytics/snapshots-export")
def export_snapshots(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    comps = session.exec(
        select(CompetitorLink).join(Product).where(Product.shop_id == shop_id)
    ).all()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["competitor_id", "product_id", "price", "currency", "fetched_at"])
    for c in comps:
        if not domain_is_live(session, domain_from_url(c.url)):
            continue
        snaps = session.exec(
            select(PriceSnapshot).where(PriceSnapshot.competitor_link_id == c.id)
        ).all()
        for s in snaps:
            w.writerow([c.id, c.product_id, s.price, s.currency, s.fetched_at.isoformat()])
    data = buf.getvalue().encode("utf-8-sig")
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="snapshots.csv"'},
    )


@router.get("/{shop_id}/insights")
def insights(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    products = session.exec(select(Product).where(Product.shop_id == shop_id)).all()
    comps = session.exec(select(CompetitorLink).join(Product).where(Product.shop_id == shop_id)).all()
    unread = session.exec(
        select(Alert).where(Alert.shop_id == shop_id, Alert.read == False)  # noqa: E712
    ).all()
    summary = (
        f"מוצרים: {len(products)}, קישורי מתחרים: {len(comps)}, התראות שלא נקראו: {len(unread)}."
    )
    return {"summary": summary}


@router.get("/{shop_id}/members", response_model=list[MemberOut])
def members(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    owner = session.get(User, shop.owner_id)
    out: list[MemberOut] = []
    if owner:
        out.append(MemberOut(user_id=owner.id, email=owner.email, role="owner"))
    rows = session.exec(select(ShopMember).where(ShopMember.shop_id == shop_id)).all()
    for m in rows:
        u = session.get(User, m.user_id)
        if u and u.id != shop.owner_id:
            out.append(MemberOut(user_id=u.id, email=u.email, role=m.role))
    return out


@router.get("/ownership-transfer/requests/incoming", response_model=list[OwnershipTransferOut])
def ownership_transfer_incoming(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    now = utcnow()
    rows = session.exec(
        select(ShopOwnershipTransfer)
        .where(ShopOwnershipTransfer.to_user_id == user.id)
        .order_by(ShopOwnershipTransfer.id.desc()),
    ).all()
    out: list[OwnershipTransferOut] = []
    for r in rows:
        if r.status == "pending" and r.expires_at <= now:
            r.status = "expired"
            r.responded_at = now
            session.add(r)
            session.commit()
            session.refresh(r)
        out.append(_ownership_transfer_to_out(session, r))
    return out


@router.get("/ownership-transfer/requests/outgoing", response_model=list[OwnershipTransferOut])
def ownership_transfer_outgoing(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    rows = session.exec(
        select(ShopOwnershipTransfer)
        .where(ShopOwnershipTransfer.from_user_id == user.id)
        .order_by(ShopOwnershipTransfer.id.desc()),
    ).all()
    return [_ownership_transfer_to_out(session, r) for r in rows]


@router.post("/{shop_id}/ownership-transfer/request", response_model=OwnershipTransferOut)
def ownership_transfer_request(
    shop_id: int,
    body: OwnershipTransferCreateIn,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    if shop.owner_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "רק בעל החנות יכול להעביר בעלות")

    target_email = body.target_email.strip().lower()
    if not target_email:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "יש להזין אימייל יעד")
    target = session.exec(select(User).where(User.email == target_email)).first()
    if not target:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "לא נמצא משתמש עם האימייל הזה")
    if target.id == user.id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "לא ניתן להעביר בעלות לעצמך")

    exists_pending = session.exec(
        select(ShopOwnershipTransfer).where(
            ShopOwnershipTransfer.shop_id == shop_id,
            ShopOwnershipTransfer.status == "pending",
        ),
    ).first()
    if exists_pending:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "כבר קיימת בקשת העברה ממתינה לחנות זו")

    row = ShopOwnershipTransfer(
        shop_id=shop_id,
        from_user_id=user.id,
        to_user_id=target.id,
        to_email=target.email,
        status="pending",
        note=(body.note or "").strip() or None,
        expires_at=utcnow() + timedelta(days=7),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return _ownership_transfer_to_out(session, row)


@router.post("/ownership-transfer/requests/{request_id}/approve", response_model=OwnershipTransferOut)
def ownership_transfer_approve(
    request_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    row = session.get(ShopOwnershipTransfer, request_id)
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "בקשת העברה לא נמצאה")
    if row.to_user_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "אפשר לאשר רק בקשה שיועדה אליך")
    if row.status != "pending":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "הבקשה כבר לא ממתינה")
    now = utcnow()
    if row.expires_at <= now:
        row.status = "expired"
        row.responded_at = now
        session.add(row)
        session.commit()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "הבקשה פגה — יש לבקש העברה חדשה")

    shop = session.get(Shop, row.shop_id)
    if not shop:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "החנות לא נמצאה")
    if shop.owner_id != row.from_user_id:
        row.status = "canceled"
        row.responded_at = now
        session.add(row)
        session.commit()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "בעלות החנות כבר השתנתה — הבקשה בוטלה")

    old_owner_id = shop.owner_id
    shop.owner_id = user.id
    session.add(shop)

    old_owner_member = session.exec(
        select(ShopMember).where(ShopMember.shop_id == shop.id, ShopMember.user_id == old_owner_id),
    ).first()
    if old_owner_member:
        old_owner_member.role = "member"
        session.add(old_owner_member)
    else:
        session.add(ShopMember(shop_id=shop.id, user_id=old_owner_id, role="member"))

    new_owner_member = session.exec(
        select(ShopMember).where(ShopMember.shop_id == shop.id, ShopMember.user_id == user.id),
    ).first()
    if new_owner_member:
        session.delete(new_owner_member)

    row.status = "accepted"
    row.responded_at = now
    session.add(row)

    pending_same_shop = session.exec(
        select(ShopOwnershipTransfer).where(
            ShopOwnershipTransfer.shop_id == shop.id,
            ShopOwnershipTransfer.status == "pending",
            ShopOwnershipTransfer.id != row.id,
        ),
    ).all()
    for p in pending_same_shop:
        p.status = "canceled"
        p.responded_at = now
        session.add(p)

    session.commit()
    session.refresh(row)
    return _ownership_transfer_to_out(session, row)


@router.post("/ownership-transfer/requests/{request_id}/decline", response_model=OwnershipTransferOut)
def ownership_transfer_decline(
    request_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    row = session.get(ShopOwnershipTransfer, request_id)
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "בקשת העברה לא נמצאה")
    if row.to_user_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "אפשר לדחות רק בקשה שיועדה אליך")
    if row.status != "pending":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "הבקשה כבר לא ממתינה")
    row.status = "declined"
    row.responded_at = utcnow()
    session.add(row)
    session.commit()
    session.refresh(row)
    return _ownership_transfer_to_out(session, row)


@router.post("/ownership-transfer/requests/{request_id}/cancel", response_model=OwnershipTransferOut)
def ownership_transfer_cancel(
    request_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    row = session.get(ShopOwnershipTransfer, request_id)
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "בקשת העברה לא נמצאה")
    if row.from_user_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "רק השולח יכול לבטל בקשת העברה")
    if row.status != "pending":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "הבקשה כבר לא ממתינה")
    row.status = "canceled"
    row.responded_at = utcnow()
    session.add(row)
    session.commit()
    session.refresh(row)
    return _ownership_transfer_to_out(session, row)


@router.post("/{shop_id}/invites")
def create_invite(
    shop_id: int,
    body: InviteCreate,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    token = secrets.token_urlsafe(24)
    inv = Invite(shop_id=shop_id, email=body.email, token=token, role=body.role)
    session.add(inv)
    session.commit()
    return {"token": token}


@router.get("/{shop_id}/api-keys", response_model=list[ApiKeyOut])
def list_api_keys(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    keys = session.exec(select(ApiKey).where(ApiKey.shop_id == shop_id)).all()
    return [
        ApiKeyOut(id=k.id, name=k.name, created_at=k.created_at, prefix=k.prefix) for k in keys
    ]


@router.post("/{shop_id}/api-keys")
def create_api_key(
    shop_id: int,
    body: ApiKeyCreate,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    raw = secrets.token_urlsafe(32)
    h = hashlib.sha256(raw.encode()).hexdigest()
    prefix = raw[:10]
    row = ApiKey(shop_id=shop_id, name=body.name, key_hash=h, prefix=prefix)
    session.add(row)
    session.commit()
    session.refresh(row)
    return {"id": row.id, "raw_key": raw}


@router.delete("/{shop_id}/api-keys/{key_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_api_key(
    shop_id: int,
    key_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    require_shop_access(session, user, shop_id)
    k = session.get(ApiKey, key_id)
    if not k or k.shop_id != shop_id:
        raise HTTPException(404, "לא נמצא")
    session.delete(k)
    session.commit()
    return Response(status_code=204)


@router.post("/{shop_id}/woocommerce")
def save_woo(
    shop_id: int,
    body: WooConfig,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
):
    shop = require_shop_access(session, user, shop_id)
    shop.woo_site_url = body.site_url.rstrip("/")
    shop.woo_consumer_key = body.consumer_key.strip()
    shop.woo_consumer_secret = body.consumer_secret.strip()
    cur = fetch_wc_store_currency(shop.woo_site_url, shop.woo_consumer_key, shop.woo_consumer_secret)
    if cur:
        shop.woo_currency = cur
    session.add(shop)
    session.commit()
    return {"ok": True, "woo_currency": shop.woo_currency}


@router.get("/{shop_id}/wordpress-plugin.zip")
def download_wordpress_plugin_zip(
    shop_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    api_base_override: str | None = Query(None, alias="api_base"),
):
    require_shop_access(session, user, shop_id)
    row = session.exec(
        select(WpConnectionToken).where(
            WpConnectionToken.shop_id == shop_id,
            WpConnectionToken.active == True,  # noqa: E712
        ),
    ).first()
    if not row:
        row = WpConnectionToken(
            token=secrets.token_urlsafe(64),
            shop_id=shop_id,
            created_by_user_id=user.id,
            active=True,
        )
        session.add(row)
        session.commit()
        session.refresh(row)
    override = (api_base_override or "").strip().rstrip("/")
    if override:
        if not (override.startswith("http://") or override.startswith("https://")):
            raise HTTPException(400, "api_base חייב להתחיל ב-http:// או https://")
        api_base = override
    else:
        api_base = resolve_public_api_base(session, request)
    zip_bytes = build_plugin_zip_bytes(api_base, row.token)
    buf = BytesIO(zip_bytes)
    fname = f"price-resolver-connect-shop-{shop_id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{shop_id}/competitors-import-template.xlsx")
def download_competitors_import_template(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    category: str | None = Query(None, description="שם קטגוריה (כמו שמופיע במוצרים)"),
):
    require_shop_access(session, user, shop_id)
    stmt = select(Product).where(Product.shop_id == shop_id)
    cat = (category or "").strip()
    if cat:
        stmt = stmt.where(Product.category_name == cat)
    products = session.exec(stmt.order_by(Product.id)).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "import"
    ws.append(["product_id", "sku", "product_name", "category", "competitor_url", "competitor_label"])
    for p in products:
        ws.append([p.id, p.sku or "", p.name, getattr(p, "category_name", None) or "", "", ""])
    ws2 = wb.create_sheet("instructions")
    ws2.append(["איך ממלאים את הקובץ"])
    ws2.append(["1) לא משנים את product_id / sku / product_name / category."])
    ws2.append(["2) ממלאים competitor_url עם קישור מלא למוצר אצל מתחרה."])
    ws2.append(["3) competitor_label אופציונלי; אם ריק, יילקח שם קיים או דומיין."])
    ws2.append(["4) אפשר כמה שורות לאותו מוצר — קישור אחד בכל שורה."])
    ws2.append(["5) שורות בלי competitor_url יידלגו ולא יעשו נזק."])
    ws2.append(["6) בסוף מעלים את הקובץ במסך מוצרים ומתחרים."])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if cat:
        safe = re.sub(r"[^A-Za-z0-9._-]+", "-", cat).strip("-").lower()
        suffix = f"-{safe[:48]}" if safe else "-category"
    else:
        suffix = "-all"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="competitors-import-template-{shop_id}{suffix}.xlsx"',
        },
    )


def _normalize_header_cell(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    return s.replace(" ", "_").replace("-", "_")


def _xlsx_row_cell(row: tuple[Any, ...] | list[Any] | None, col_map: dict[str, int], name: str) -> str | None:
    idx = col_map.get(name)
    if idx is None or row is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    return str(v).strip()


@router.post("/{shop_id}/competitors-import")
async def import_competitors_from_xlsx(
    shop_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    require_shop_access(session, user, shop_id)
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "קובץ Excel לא תקין") from None
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "הקובץ ריק") from None
    col_map: dict[str, int] = {}
    aliases = {
        "product_id": "product_id",
        "מזהה_מוצר": "product_id",
        "id": "product_id",
        "sku": "sku",
        "מקט": "sku",
        "product_name": "product_name",
        "שם_מוצר": "product_name",
        "name": "product_name",
        "competitor_url": "competitor_url",
        "קישור_למתחרה": "competitor_url",
        "url": "competitor_url",
        "competitor_label": "competitor_label",
        "שם_מתחרה": "competitor_label",
        "label": "competitor_label",
    }
    for i, cell in enumerate(header_row):
        key = _normalize_header_cell(cell)
        canon = aliases.get(key, key)
        if canon in ("product_id", "sku", "product_name", "competitor_url", "competitor_label"):
            col_map[canon] = i
    if "product_id" not in col_map or "competitor_url" not in col_map:
        raise HTTPException(
            400,
            "חסרות עמודות חובה: product_id ו-competitor_url (שורת כותרת ראשונה)",
        )

    added = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows_iter, start=2):
        if not row:
            continue
        pid_raw = _xlsx_row_cell(tuple(row), col_map, "product_id")
        url_raw = _xlsx_row_cell(tuple(row), col_map, "competitor_url") or ""
        if not url_raw:
            skipped += 1
            continue
        try:
            pid = int(float(str(pid_raw).replace(",", "").strip()))
        except (TypeError, ValueError):
            errors.append(f"שורה {row_idx}: מזהה מוצר לא תקף")
            skipped += 1
            continue

        p = session.get(Product, pid)
        if not p or p.shop_id != shop_id:
            errors.append(f"שורה {row_idx}: מוצר {pid} לא שייך לחנות")
            skipped += 1
            continue

        url = _normalize_competitor_url(url_raw)
        if not url:
            errors.append(f"שורה {row_idx}: קישור לא תקף")
            skipped += 1
            continue

        lab_raw = _xlsx_row_cell(tuple(row), col_map, "competitor_label")
        lab = lab_raw if lab_raw else None

        existing = session.exec(
            select(CompetitorLink).where(
                CompetitorLink.product_id == pid,
                CompetitorLink.url == url,
            ),
        ).first()
        if existing:
            if lab is not None and (existing.label or "") != (lab or ""):
                existing.label = lab
                tid = getattr(existing, "tracked_competitor_id", None)
                if tid and lab:
                    tc = session.get(TrackedCompetitor, tid)
                    if tc:
                        tc.display_name = lab[:240]
                        session.add(tc)
                session.add(existing)
                updated += 1
            else:
                skipped += 1
            continue

        tc = _resolve_tracked_competitor(
            session,
            shop_id,
            url,
            None,
            lab,
            lab,
            allow_domain_fallback_name=True,
        )
        c = CompetitorLink(
            product_id=pid,
            url=url,
            label=tc.display_name,
            tracked_competitor_id=tc.id,
        )
        session.add(c)
        session.flush()
        ensure_domain_review_queue_item_for_competitor(session, shop_id, c.id, try_fetch=False)
        added += 1

    session.commit()
    return {"added": added, "updated": updated, "skipped": skipped, "errors": errors[:50]}
